from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import string
import re
import sys

def findMetric(ws):
    # Find the column with metric types
    for row in ws.iter_rows(min_col = 1, min_row = tablerow, max_col= columnCount, max_row = tablerow):
        for cell in row:
            if cell.value == 'Metric_Type':
                metricColumn = cell.column
                return metricColumn


def removeMetrics(check, ws):
    doomlist = []

    # Find the column with metric types
    metricColumn = findMetric(ws)

    # Go through each row to check its metric type
    for row in ws.iter_rows(min_col = metricColumn, min_row=tablerow, max_col = metricColumn, max_row = rowCount):
        for cell in row:
            if cell.value == check:
                doomlist.append(cell.row)
    
    # Destroy the rows we don't want. 
    for row in reversed(doomlist):
        ws.delete_rows(row)


def sheetsplit(ws):
    print('sheetsplit!')
    wb.copy_worksheet(ws)
    if wbtype == 'TR_B1':
        ws.title = 'TR B1 Unique COUNTER 5'
    else:        
        ws.title = 'TR J1 Unique COUNTER 5'
    removeMetrics('Total_Item_Requests', ws)

    # switch to other
    ws = wb['Sheet1 Copy']
    if wbtype == 'TR_B1':
        ws.title = 'TR B1 Total COUNTER 5'
        removeMetrics('Unique_Title_Requests', ws)
    else:
        ws.title = 'TR J1 Total COUNTER 5'
        removeMetrics('Unique_Item_Requests', ws)


def tablesplit(ws):
    print('tablesplit!')
    metricColumn = findMetric(ws)
    for row in ws.iter_rows(min_col = 1, min_row = tablerow, max_col = columnCount, max_row = rowCount):
        if row[metricColumn].value == 'Unique_Item_Requests':
            continue
        else:
            for cell in row: 
                col = cell.column
                col = string.ascii_uppercase[col]
                newrow = cell.row - tablerow + rowCount
                ws[f'{col}{newrow}'] = cell.value
            if row[metricColumn].value == 'Total_Item_Requests':
                ws.delete_rows(cell.row)


# BEGIN

if len(sys.argv) != 3:
   print(sys.argv)
   sys.exit("Please include the name of the file being tweaked and a file name where the output should go")

wb = load_workbook(sys.argv[1])

# wb = load_workbook('TR_B1_Input.xlsx')
ws = wb.active
wbtype = ws['B2'].value

# Initial styles

for row in ws.iter_rows(min_col = 1, min_row = 1, max_col = ws.max_column, max_row = ws.max_row):
    for cell in row:
        cell.font = Font(name = 'Calibri', size = 12)



# METADATA SECTION

metadataKeeps = ['Report_Name', 'Report_ID', 'Reporting_Period', 'Created']
metadataGoodbyes = []

# Find the rows we don't want
for count, row in enumerate(ws.rows, 1):
    # Reach the end of the metadata section
    
    if row[0].value == None:
        break
    for cell in row:
        cell.font = Font(name = 'Calibri', size = 12, bold=True)
    if row[0].value not in metadataKeeps:
        metadataGoodbyes.append(count)

# Purge the rows we don't want
for row in reversed(metadataGoodbyes):
    ws.delete_rows(row)

# Reformat the dates in the remaining rows.
rp = ws['B3'].value
rp = re.sub(r'[\w_]+?=', '', rp)
rp = re.sub(';', ' to', rp)
ws['B3'].value = rp

cr = ws['B4'].value
cr = re.sub('T.*', '', cr)
ws['B4'].value = cr


#DATA SECTION

rowCount = ws.max_row

#keep the metadata that's going to be wiped when columns are removed, reinsert later...
keepMetadatas = []
for Bcell in ws.iter_rows(min_col=2, min_row=1, max_col=2, max_row=4, values_only=True):
    Bcell = ''.join(Bcell)
    keepMetadatas.append(Bcell)



# Columns to be removed. 
dataGoodbyes = ['Publisher', 'Publisher_ID', 'DOI', 'Proprietary_ID', 'URI']

if wbtype == 'TR_B1':
    dataGoodbyes.append('Print_ISSN')
    dataGoodbyes.append('Online_ISSN')

dataGoodbyeColumns = []


# find header row of table to be used through rest of program as variable
tablerow = ''
for count, row in enumerate(ws.iter_rows(min_col=1, min_row=1, max_col=1, max_row=25, values_only=True), 1):
    for cell in row:
        if cell == 'Title':
            tablerow = count
            break



# find the table columns we don't want
columnCount = ws.max_column
for count, column in enumerate(ws.iter_cols(min_row=tablerow, max_col=columnCount, max_row=tablerow, values_only=True), 1):
    # Reach the end of the metadata section
    try: 
        column = ''.join(column)
    except:
        break
    if column in dataGoodbyes:
        dataGoodbyeColumns.append(count)

# Purge the columns we don't want
for column in reversed(dataGoodbyeColumns):
    ws.delete_cols(column)

# reinsert metadata
for row in ws.iter_rows(min_col=2, min_row=1, max_col=2, max_row=4):
    for cell in row: 
        cell.value = keepMetadatas.pop(0)
        cell.font = Font(name = 'Calibri', bold = True)

# color metadata
for row in ws.iter_rows(min_col = 1, min_row = 1, max_col = 5, max_row = 3):
    for cell in row:
        if cell.row == 3:
            cell.fill = PatternFill(start_color='FF9999', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='CCFF99', fill_type='solid')

# find the 'total' column and move it to the end. Along the way, change the titles of columns. 
origcolumn = ''
movecolumn = ''
breaker = 0

columnCount = ws.max_column + 1 # + 1 to give space for the reinsertion
for count, column in enumerate(ws.iter_cols(min_col=1, min_row=tablerow, max_col=columnCount, max_row=tablerow, values_only=True), 1):
    for cell in column:
        if cell == 'Reporting_Period_Total':
            origcolumn = count
            columnletter = string.ascii_uppercase[origcolumn - 1]
        if cell == None:
            movement = count - origcolumn
            ws.move_range(f'{columnletter}1:{columnletter}{rowCount}', cols = movement)
            
            # Delete the original column 
            ws.delete_cols(origcolumn)
            breaker = 1
            break
    if breaker == 1:
        break


# henceforth, this is how wide this table is! 
columnCount = ws.max_column

# switch headers to what we like
for column in ws.iter_cols(min_col = 1, min_row = tablerow, max_col = columnCount, max_row = tablerow):
    for cell in column:
        cell.value = cell.value.replace('-2020', '')
        cell.value = cell.value.replace('Reporting_Period_Total', 'YTD Total')
        cell.font = Font(color = '000000', name = 'Calibri', bold=True, size = 12)
        cell.fill = PatternFill(start_color='99CCFF', fill_type='solid')


# add row of sums! 
ws.insert_rows(tablerow + 1)
sumRow = ws.row_dimensions[tablerow + 1]
sumRow.font = Font(name = 'Calibri', size = 12, bold = True)
sumRow.fill = PatternFill(start_color='99CCFF', fill_type='solid')

rowCount = ws.max_row # henceforth, this is how tall the table is!

for count, column in enumerate(ws.iter_cols(min_col = origcolumn, min_row = tablerow + 1, max_col = columnCount, max_row = tablerow), 5):
    columnletter = string.ascii_uppercase[count]
    cellcode = f'{columnletter}{tablerow + 1}'
    ws[cellcode].value = f"=SUM({columnletter}{tablerow + 2}:{columnletter}{rowCount})"
    ws[cellcode].font = Font(name = 'Calibri', bold = True, size = 12)
    ws[cellcode].fill = PatternFill(start_color='99CCFF', fill_type='solid')

# add 'Total' to first cell in that row
ws[f'A{tablerow + 1}'].value = 'Total'
ws[f'A{tablerow + 1}'].font = Font(name = 'Calibri', bold = True, size = 12)
ws[f'A{tablerow + 1}'].fill = PatternFill(start_color='99CCFF', fill_type='solid')

# separate totals and uniques
if (rowCount - tablerow) > 20:
    sheetsplit(ws)
else: 
    tablesplit(ws)


# END

wb.save(sys.argv[2])
# wb.save('Output.xlsx')